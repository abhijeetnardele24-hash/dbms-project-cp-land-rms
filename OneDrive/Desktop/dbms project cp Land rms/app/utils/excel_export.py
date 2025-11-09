"""
Excel Export utilities for generating spreadsheet reports.
Uses openpyxl library for Excel file generation.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


def generate_properties_excel(properties):
    """
    Generate Excel report of properties.
    
    Args:
        properties: List of Property objects
        
    Returns:
        BytesIO: Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Properties"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['ULPIN', 'Property Type', 'District', 'Locality', 'Pincode', 
               'Area (sqft)', 'Market Value (₹)', 'Owner', 'Created Date']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, prop in enumerate(properties, 2):
        ws.cell(row=row_num, column=1, value=prop.ulpin)
        ws.cell(row=row_num, column=2, value=prop.property_type.replace('_', ' ').title())
        ws.cell(row=row_num, column=3, value=prop.district)
        ws.cell(row=row_num, column=4, value=prop.locality)
        ws.cell(row=row_num, column=5, value=prop.pincode)
        ws.cell(row=row_num, column=6, value=prop.area_sqft)
        ws.cell(row=row_num, column=7, value=prop.market_value)
        ws.cell(row=row_num, column=8, value=prop.owner.full_name if prop.owner else 'N/A')
        ws.cell(row=row_num, column=9, value=prop.created_at.strftime('%Y-%m-%d'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_mutations_excel(mutations):
    """
    Generate Excel report of mutations.
    
    Args:
        mutations: List of Mutation objects
        
    Returns:
        BytesIO: Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mutations"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Mutation Number', 'Certificate Number', 'Property ULPIN', 'Mutation Type', 
               'Previous Owner', 'New Owner', 'Status', 'Application Date', 'Approval Date']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, mut in enumerate(mutations, 2):
        ws.cell(row=row_num, column=1, value=mut.mutation_number or 'N/A')
        ws.cell(row=row_num, column=2, value=mut.mutation_certificate_number or 'N/A')
        ws.cell(row=row_num, column=3, value=mut.property.ulpin)
        ws.cell(row=row_num, column=4, value=mut.mutation_type.replace('_', ' ').title())
        ws.cell(row=row_num, column=5, value=mut.previous_owners or 'N/A')
        ws.cell(row=row_num, column=6, value=mut.new_owners or 'N/A')
        ws.cell(row=row_num, column=7, value=mut.status.replace('_', ' ').title())
        ws.cell(row=row_num, column=8, value=mut.created_at.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=9, value=mut.approval_date.strftime('%Y-%m-%d') if mut.approval_date else 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_payments_excel(payments):
    """Generate Excel report of payments."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ['Transaction ID', 'User', 'Property ULPIN', 'Payment Type', 
               'Amount (₹)', 'Status', 'Payment Date']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1, value=payment.transaction_id)
        ws.cell(row=row_num, column=2, value=payment.user.full_name)
        ws.cell(row=row_num, column=3, value=payment.property.ulpin if payment.property else 'N/A')
        ws.cell(row=row_num, column=4, value=payment.payment_type.replace('_', ' ').title())
        ws.cell(row=row_num, column=5, value=payment.amount)
        ws.cell(row=row_num, column=6, value=payment.status.title())
        ws.cell(row=row_num, column=7, value=payment.created_at.strftime('%Y-%m-%d'))
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_users_excel(users):
    """Generate Excel report of users."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ['Name', 'Email', 'Phone', 'Role', 'Status', 'Created Date', 'Last Login']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1, value=user.full_name)
        ws.cell(row=row_num, column=2, value=user.email)
        ws.cell(row=row_num, column=3, value=user.phone or 'N/A')
        ws.cell(row=row_num, column=4, value=user.role.title())
        ws.cell(row=row_num, column=5, value='Active' if user.is_active else 'Inactive')
        ws.cell(row=row_num, column=6, value=user.created_at.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=7, value=user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never')
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
